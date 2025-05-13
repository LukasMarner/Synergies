# Run power flow & results

import pandapower as pp
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from Data_preprocessing_BH import *
import create_BH_net
from create_BH_net import *
import importlib
importlib.reload(create_BH_net)

def check_power_system_violations(net, print_results = True):
    """
    Checks for voltage violations, overloaded transformers, and overloaded lines in the power system.
    Prints the results.
    """
    pp.runpp(net)
    # Add corresponding bus name to output
    net.res_bus['Bus name'] = net.bus['name']
    
    # Check for voltage violations
    voltage_violations = net.res_bus[(net.res_bus.vm_pu > net.bus.max_vm_pu) | (net.res_bus.vm_pu < net.bus.min_vm_pu)]
        # Check if index 28 exists before dropping it
    if 28 in voltage_violations.index:
        voltage_violations = voltage_violations.drop(28)  # remove ref bus
    # voltage_violations= voltage_violations.drop(28)  # remove ref bus
    
    if not voltage_violations.empty:
        if print_results:
            print("Voltage violations detected at the following buses:")
        v_voltages = True
    else:
        voltage_violations = 'No voltage violations detected.'
        v_voltages = False
        
    if print_results:
        print(voltage_violations)
    
    # Check for overloaded transformers
    net.res_trafo['name'] = net.trafo['name']
    overloaded_transformers = net.res_trafo[net.res_trafo['loading_percent'] > 100]
    
    if not overloaded_transformers.empty:
        if print_results:
            print("Overloaded transformers detected:")
        v_trafos = True
    else:
        overloaded_transformers = 'No overloaded transformers detected.'
        v_trafos = False
    if print_results:
        print(overloaded_transformers)


    trafo_loadings = net.res_trafo[['name', 'loading_percent']]
    
    # Add bus names to line data
    net.line['from_bus_name'] = net.line['from_bus'].map(net.bus['name'])
    net.line['to_bus_name'] = net.line['to_bus'].map(net.bus['name'])
    
    lines = net.line[['from_bus', 'to_bus', 'from_bus_name', 'to_bus_name']]
    line_loadings = net.res_line['loading_percent']
    line_loadings = pd.concat([lines, line_loadings], axis=1)

    # Check for overloaded lines
    overloaded_lines = line_loadings[line_loadings.loading_percent > 100]

    if not overloaded_lines.empty:
        if print_results:
            print("Overloaded lines detected:")
        v_lines = True
    else:
        overloaded_lines = "No overloaded lines detected."
        v_lines = False
    
    if print_results:
        print(overloaded_lines)
        print('Line loadings:\n', line_loadings)
        print('Bus voltages and power injections:\n', net.res_bus)
        print('Transformer loadings:\n', trafo_loadings)


    bus_voltages_injections = net.res_bus

    
    # if overloaded_lines == "No overloaded lines detected." & overloaded_transformers == 'No overloaded transformers detected.'  & voltage_violations == 'No voltage violations detected.':
    #     secure = True
    # else:
    #     secure = False
    violations_df = pd.DataFrame({
    'overloaded_lines': [v_lines],
    'overloaded_transformers': [v_trafos],
    'voltage_violations': [v_voltages]})

    # if (overloaded_lines[0] == "No overloaded lines detected." and 
    #     overloaded_transformers == 'No overloaded transformers detected.' and 
    #     voltage_violations == 'No voltage violations detected.'):
    #     secure = True
    # else:
    #     secure = False
    # print(violations_df)
    if ((violations_df == False).all().all()):
        secure = True
    else:
        secure = False
    if print_results:
        print('Sytem is secure:', secure)

    return secure, voltage_violations, overloaded_lines, overloaded_transformers, line_loadings, trafo_loadings, bus_voltages_injections


# Example usage:
# net = pp.create_some_example_network()  # Create or load your pandapower network
# check_power_system_violations(net)

import pandas as pd
import pandapower as pp  # Assuming pandapower is used
import numpy as np

def perform_power_flow_analysis(measurement_path, new_path, N_substations=16, measurement_type = 'smart meter'):
    # Initialize lists to store the results
    datetimes = []
    secure_list = []
    voltage_violations_list = []
    overloaded_lines_list = []
    overloaded_transformers_list = []
    line_loadings_list = []
    trafo_loadings_list = []
    bus_voltages_injections_list = []

    # Read the CSV file
    measurement_csv = pd.read_csv(measurement_path)

    for i in range(int(len(measurement_csv) / N_substations)):
        # Get the subset for the current iteration
        measurement_data = measurement_csv[i * N_substations:(i + 1) * N_substations]

        # Extract the datetime for the current subset
        current_datetime = measurement_data['datetime'].iloc[0]
        datetimes.append(current_datetime)

        # Initialize and run power system analysis
        net = pp.create_empty_network()
        net = load_data(new_path, net)
        if measurement_type == 'smart meter':
            net = net_60kV_measurements(net, measurement_data)
        elif measurement_type == 'scada':
            net = net_60kV_SCADA_measurements(net, measurement_data)
        else:
            print('Invalid measurement type')
            
        # pp.runpp(net)

        # Check power system violations
        secure, voltage_violations, overloaded_lines, overloaded_transformers, line_loadings, trafo_loadings, bus_voltages_injections = check_power_system_violations(net)

        # Store results
        secure_list.append(secure)
        
        # Ensure the violation results are in DataFrame format
        if isinstance(voltage_violations, str):
            voltage_violations = pd.DataFrame(columns=['vm_pu', 'va_degree', 'p_mw', 'q_mvar', 'Bus name'])
        if isinstance(overloaded_lines, str):
            overloaded_lines = pd.DataFrame(columns=['line_name', 'line_loading_percent'])
        if isinstance(overloaded_transformers, str):
            overloaded_transformers = pd.DataFrame(columns=['trafo_name', 'trafo_loading_percent'])
        if isinstance(line_loadings, str):
            line_loadings = pd.DataFrame(columns=['line_name', 'line_loading_percent'])
        if isinstance(trafo_loadings, str):
            trafo_loadings = pd.DataFrame(columns=['trafo_name', 'trafo_loading_percent'])
        if isinstance(bus_voltages_injections, str):
            bus_voltages_injections = pd.DataFrame(columns=['bus_name', 'voltage_pu'])

        voltage_violations_list.append(voltage_violations)
        overloaded_lines_list.append(overloaded_lines)
        overloaded_transformers_list.append(overloaded_transformers)
        line_loadings_list.append(line_loadings)
        trafo_loadings_list.append(trafo_loadings)
        bus_voltages_injections_list.append(bus_voltages_injections)
    
    return datetimes, secure_list, voltage_violations_list, overloaded_lines_list, overloaded_transformers_list, line_loadings_list, trafo_loadings_list, bus_voltages_injections_list

from openpyxl import load_workbook

def write_results_to_excel(excel_path, datetimes, secure_list, voltage_violations_list, overloaded_lines_list, overloaded_transformers_list, line_loadings_list, trafo_loadings_list, bus_voltages_injections_list):
    # Create DataFrames for each result type with Datetime
    df_secure = pd.DataFrame({
        'Datetime': datetimes,
        'Secure': secure_list
    })

    # Flatten the lists of nested DataFrames and include the datetime
    df_voltage_violations = pd.concat([df.assign(Datetime=dt) for df, dt in zip(voltage_violations_list, datetimes)], ignore_index=True)
    df_overloaded_lines = pd.concat([df.assign(Datetime=dt) for df, dt in zip(overloaded_lines_list, datetimes)], ignore_index=True)
    df_overloaded_transformers = pd.concat([df.assign(Datetime=dt) for df, dt in zip(overloaded_transformers_list, datetimes)], ignore_index=True)
    df_line_loadings = pd.concat([df.assign(Datetime=dt) for df, dt in zip(line_loadings_list, datetimes)], ignore_index=True)
    df_trafo_loadings = pd.concat([df.assign(Datetime=dt) for df, dt in zip(trafo_loadings_list, datetimes)], ignore_index=True)
    df_bus_voltages_injections = pd.concat([df.assign(Datetime=dt) for df, dt in zip(bus_voltages_injections_list, datetimes)], ignore_index=True)

    # Write DataFrames to different sheets
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df_secure.to_excel(writer, sheet_name='Secure', index=False)
        df_voltage_violations.to_excel(writer, sheet_name='Voltage Violations', index=False)
        df_overloaded_lines.to_excel(writer, sheet_name='Overloaded Lines', index=False)
        df_overloaded_transformers.to_excel(writer, sheet_name='Overloaded Transformers', index=False)
        df_line_loadings.to_excel(writer, sheet_name='Line Loadings', index=False)
        df_trafo_loadings.to_excel(writer, sheet_name='Transformer Loadings', index=False)
        df_bus_voltages_injections.to_excel(writer, sheet_name='Bus Voltages Injections', index=False)

    # Open the workbook to add filters
    workbook = load_workbook(excel_path)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        # Add a filter to the first row
        sheet.auto_filter.ref = sheet.dimensions

    # Save the workbook with the filters applied
    workbook.save(excel_path)

def write_results_to_csv(excel_path):
    # Path to the Excel file

    # Read the Excel file
    excel_data = pd.read_excel(excel_path, sheet_name=None)  # sheet_name=None reads all sheets

    # Iterate over each sheet
    for sheet_name, data in excel_data.items():
        # Define the CSV file name
        csv_file_path = f'{sheet_name}.csv'
    
        # Write the sheet data to a CSV file
        data.to_csv(csv_file_path, index=False)

        



def check_power_system_violations_contingency(net):
    """
    Checks for voltage violations, overloaded transformers, and overloaded lines in the power system.
    Prints the results.
    """
    net.res_bus['Bus name'] = net.bus['name']
    
    # Check for voltage violations
    voltage_violations = net.res_bus[(net.res_bus.vm_pu > net.bus.max_vm_pu) | (net.res_bus.vm_pu < net.bus.min_vm_pu)]
    voltage_violations = voltage_violations.drop(index=28)  # remove ref bus
    violations = {}
    
    if not voltage_violations.empty:
        violations['voltage violations']

    else:
        print("No voltage violations detected.")
    
    # Check for overloaded transformers
    overloaded_transformers = net.res_trafo[net.res_trafo.loading_percent > 100]
    
    if not overloaded_transformers.empty:
        print("Overloaded transformers detected:")
        print(overloaded_transformers)
    else:
        print("No overloaded transformers detected.")
    
    # Check for overloaded lines
    overloaded_lines = net.res_line[net.res_line.loading_percent > 100]
    
    if not overloaded_lines.empty:
        print("Overloaded lines detected:")
        print(overloaded_lines)
    else:
        print("No overloaded lines detected.")
    
    # Add bus names to line data
    net.line['from_bus_name'] = net.line['from_bus'].map(net.bus['name'])
    net.line['to_bus_name'] = net.line['to_bus'].map(net.bus['name'])
    
    lines = net.line[['from_bus', 'to_bus', 'from_bus_name', 'to_bus_name']]
    line_loadings = net.res_line['loading_percent']
    line_loadings = pd.concat([lines, line_loadings], axis=1)
    print('Line loadings:\n', line_loadings)
    
    print('Bus voltages and power injections:\n', net.res_bus)
    # Add corresponding bus name to output
    # if (not overloaded_transformers.empty) & (not voltage_violations.empty) & (not overloaded_lines.empty):
    #     print('Critical ')




def visualize_voltage_distributions(net, VaR = None, CVaR = None):
    # Extract voltage magnitudes
    voltage_magnitudes = net.res_bus.vm_pu.values

    # Plot voltage distribution
    plt.figure(figsize=(8, 6))
    plt.hist(voltage_magnitudes, bins=int(len(net.bus)/5), color='skyblue', edgecolor='black')
    plt.xlabel('Voltage Magnitude (pu)')
    plt.ylabel('Frequency')
    plt.title('Voltage Magnitude Distribution')
    plt.grid(True)

    if VaR is not None:
        plt.axvline(x=VaR[0], color = 'r', linestyle ='--', label ='VaR lower bound')
        plt.axvline(x=VaR[1], color = 'r', linestyle ='-', label ='VaR upper bound')
    if CVaR is not None:
        plt.axvline(x=CVaR[0], color = 'black', linestyle ='--', label ='CVaR lower bound')
        plt.axvline(x=CVaR[1], color = 'black', linestyle ='-', label ='CVaR lower bound')
    plt.legend()
    plt.show()

def visualize_lineloadings_distributions(net, VaR = None, CVaR = None):
    # Extract voltage magnitudes
    line_loadings = net.res_line.loading_percent

    # Plot voltage distribution
    plt.figure(figsize=(8, 6))
    plt.hist(line_loadings, bins=int(len(net.line)/3), color='skyblue', edgecolor='black')
    plt.xlabel('Line loading (%)')
    plt.ylabel('Frequency')
    plt.title('Line Loadings Distribution')
    if VaR is not None:
        plt.axvline(x=VaR, color = 'r', linestyle ='-', label ='VaR')
    if CVaR is not None:
        plt.axvline(x=CVaR, color = 'black', linestyle ='-', label ='CVaR')
    
    plt.legend()

    plt.grid(True)
    plt.show()

def visualize_trafoloadings_distributions(net, VaR=None, CVaR=None):
    # Extract voltage magnitudes
    trafo_loadings = net.res_trafo.loading_percent

    # Plot voltage distribution
    plt.figure(figsize=(8, 6))
    plt.hist(trafo_loadings, bins=int(len(net.trafo)/3), color='skyblue', edgecolor='black')
    plt.xlabel('Transformer loading (%)')
    plt.ylabel('Frequency')
    plt.title('Transformer Loadings Distribution')

    if VaR is not None:
        plt.axvline(x=VaR, color = 'r', linestyle ='-', label ='VaR')
    if CVaR is not None:
        plt.axvline(x=CVaR, color = 'black', linestyle ='-', label ='CVaR')

    plt.legend()
    plt.grid(True)
    plt.show()


def compute_var_cvar(array, confidence_level=0.95):
    # Sort the array of line loadings in ascending order
    array = array.dropna()
    # nan_mask = np.isnan(array)
    # array = array[nan_mask]
    sorted_array = np.sort(array)
    
    # Determine the index corresponding to the desired percentile
    index = int(len(sorted_array) * confidence_level)
    
    # Compute the VaR
    var = sorted_array[index]
    cvar = np.mean(sorted_array[index:])
    
    return var, cvar

def compute_var_voltages(voltage_magnitudes, confidence_level=0.95):
    """
    Calculate the Value at Risk (VaR) for voltage magnitudes.

    Parameters:
    voltage_magnitudes (array-like): Array of voltage magnitudes.
    confidence_level (float): Confidence level for VaR.

    Returns:
    tuple: Lower and upper bounds of the VaR interval.
    """
    voltage_magnitudes = voltage_magnitudes.dropna()
    # Calculate the lower and upper percentiles
    lower_percentile = (1 - confidence_level) / 2
    upper_percentile = 1 - lower_percentile

    # Compute the VaR interval
    lower_bound = np.percentile(voltage_magnitudes, lower_percentile * 100)
    upper_bound = np.percentile(voltage_magnitudes, upper_percentile * 100)

    var = [lower_bound, upper_bound]

    sorted_voltages = np.sort(voltage_magnitudes)
    
    # Calculate the index for the VaR
    var_index = int(len(sorted_voltages) * (confidence_level))
    
    # Calculate the CVaR as the mean of the values below the VaR
    cvar_values = sorted_voltages[:var_index]
    cvar = np.mean(cvar_values)
    
    # Calculate CVaR interval (minimum and maximum of the extreme values)
    cvar_lower_bound = np.min(cvar_values)
    cvar_upper_bound = np.max(cvar_values)

    cvar = [cvar_lower_bound, cvar_upper_bound]
    
    return var, cvar