# Contingency
import pandapower as pp
import Power_flow
from Power_flow import *
import numpy as np
import pandas as pd

def remove_and_run_powerflow(net, element, element_idx, element_type):
    """
    Removes an element from the network and runs a power flow analysis.
    """
    original_net = net.deepcopy()
    
    if element_type == 'line':
        pp.drop_lines(net, [element_idx])
    elif element_type == 'gen':
        net.gen.drop(index=element_idx, inplace=True)
    elif element_type == 'trafo':
        pp.drop_trafos(net, [element_idx])
    
    try:
        pp.runpp(net, numba=False)
        # print(f"Power flow successful after removing {element_type} {element_idx}.")

    except:
        print(f"Power flow did not converge after removing {element_type} {element_idx}.")
    
    return net.deepcopy()

def contingency_analysis(net):
    results = {}
    
    # Analyze lines
    for i in net.line.index:
        results[f'remove_line_{i}'] = remove_and_run_powerflow(net, net.line, i, 'line')
    
    # Analyze generators
    for i in net.gen.index:
        results[f'remove_gen_{i}'] = remove_and_run_powerflow(net, net.gen, i, 'gen')
    
    # Analyze transformers
    for i in net.trafo.index:
        results[f'remove_trafo_{i}'] = remove_and_run_powerflow(net, net.trafo, i, 'trafo')
    
    return results


def check_violations(net):
    # violations = []

    # Check transformer loadings
    trafo_loading = net.res_trafo['loading_percent']
    violated_trafos = trafo_loading[trafo_loading > 1].index
    # violations.extend([(i, 'trafo') for i in violated_trafos])

    # Check line loadings
    line_loading = net.res_line['loading_percent']
    violated_lines = line_loading[line_loading > 1].index
    # violations.extend([(i, 'line') for i in violated_lines])

    # Check bus voltages
    # Assuming net.res_bus['vm_pu'] is a pandas Series
    vm_pu_list = net.res_bus['vm_pu'].tolist()
    vm_pu_list.pop(28)

    # Convert back to a pandas Series
    net.res_bus['vm_pu'] = pd.Series(vm_pu_list)

    # Now find the minimum voltage excluding the element at index 28
    voltage_pu = net.res_bus['vm_pu']
    # min_voltage = net.res_bus['vm_pu'].min()
    # max_voltage = net.res_bus['vm_pu'].max()
    out_of_range_voltages = voltage_pu[(voltage_pu < 0.95) | (voltage_pu > 1.05)]
    if out_of_range_voltages.empty and violated_lines.empty and violated_trafos.empty:
        violations = False
    else: 
        violations = True
    # if voltage_pu < 0.95 or voltage_pu > 1.05 :
    #     violations.append(('voltage'))
    
    return violations

# def contingency_analysis_with_violations(net, measurement_prod_cons):
#     # results = []
#     outages = []
#     # print('net index', net.line.index)
#     voltage_violations_list = []
#     overloaded_lines_list = []
#     overloaded_transformers_list = []
#     line_loadings_list = []
#     trafo_loadings_list = []
#     bus_voltages_injections_list = []

#     # Analyze lines
#     for i in net.line.index:
#         net = remove_and_run_powerflow(net, net.line, i, 'line')
#         violations = check_violations(net)
#         # print('violations:',violations)
#         if violations is True:
#             outages.append(f'outage_line_{i}')
#             secure, voltage_violations, overloaded_lines, overloaded_transformers, line_loadings, trafo_loadings, bus_voltages_injections = check_power_system_violations(net, print=False)
        
#             # Ensure the violation results are in DataFrame format
#             if isinstance(voltage_violations, str):
#                 voltage_violations = pd.DataFrame(columns=['vm_pu', 'va_degree', 'p_mw', 'q_mvar', 'Bus name'])
#             if isinstance(overloaded_lines, str):
#                 overloaded_lines = pd.DataFrame(columns=['line_name', 'line_loading_percent'])
#             if isinstance(overloaded_transformers, str):
#                 overloaded_transformers = pd.DataFrame(columns=['trafo_name', 'trafo_loading_percent'])
#             if isinstance(line_loadings, str):
#                 line_loadings = pd.DataFrame(columns=['line_name', 'line_loading_percent'])
#             if isinstance(trafo_loadings, str):
#                 trafo_loadings = pd.DataFrame(columns=['trafo_name', 'trafo_loading_percent'])
#             if isinstance(bus_voltages_injections, str):
#                 bus_voltages_injections = pd.DataFrame(columns=['bus_name', 'voltage_pu'])

#             voltage_violations_list.append(voltage_violations)
#             overloaded_lines_list.append(overloaded_lines)
#             overloaded_transformers_list.append(overloaded_transformers)
#             line_loadings_list.append(line_loadings)
#             trafo_loadings_list.append(trafo_loadings)
#             bus_voltages_injections_list.append(bus_voltages_injections)
#         # results[f'check_violations(net)
#         # # results[f'remove_line_{i}'].violations = violations

#     # Analyze generators
#     for i in net.gen.index:
#         net = remove_and_run_powerflow(net, net.gen, i, 'gen')
#         violations = check_violations(net)
#         # results[f'remove_gen_{i}'] = remove_and_run_powerflow(net, net.gen, i, 'gen')
#         # violations = check_violations(results[f'remove_gen_{i}'])
#         if violations is True:
#             outages.append(f'outage_gen_{i}')
#             secure, voltage_violations, overloaded_lines, overloaded_transformers, line_loadings, trafo_loadings, bus_voltages_injections = check_power_system_violations(net, print=False)
#                         # Ensure the violation results are in DataFrame format
#             if isinstance(voltage_violations, str):
#                 voltage_violations = pd.DataFrame(columns=['vm_pu', 'va_degree', 'p_mw', 'q_mvar', 'Bus name'])
#             if isinstance(overloaded_lines, str):
#                 overloaded_lines = pd.DataFrame(columns=['line_name', 'line_loading_percent'])
#             if isinstance(overloaded_transformers, str):
#                 overloaded_transformers = pd.DataFrame(columns=['trafo_name', 'trafo_loading_percent'])
#             if isinstance(line_loadings, str):
#                 line_loadings = pd.DataFrame(columns=['line_name', 'line_loading_percent'])
#             if isinstance(trafo_loadings, str):
#                 trafo_loadings = pd.DataFrame(columns=['trafo_name', 'trafo_loading_percent'])
#             if isinstance(bus_voltages_injections, str):
#                 bus_voltages_injections = pd.DataFrame(columns=['bus_name', 'voltage_pu'])

#             voltage_violations_list.append(voltage_violations)
#             overloaded_lines_list.append(overloaded_lines)
#             overloaded_transformers_list.append(overloaded_transformers)
#             line_loadings_list.append(line_loadings)
#             trafo_loadings_list.append(trafo_loadings)
#             bus_voltages_injections_list.append(bus_voltages_injections)
#         # violations = check_violations(results[f'remove_gen_{i}'])
#         # results[f'remove_gen_{i}'].violations = violations

#     # Analyze transformers
#     for i in net.trafo.index:
#         # results[f'remove_trafo_{i}'] = remove_and_run_powerflow(net, net.trafo, i, 'trafo')
#         # violations = check_violations(results[f'remove_trafo_{i}'])
#         net = remove_and_run_powerflow(net, net.trafo, i, 'trafo')
#         violations = check_violations(net)
#         if violations is True:
#             outages.append(f'outage_trafo_{i}')
#             secure, voltage_violations, overloaded_lines, overloaded_transformers, line_loadings, trafo_loadings, bus_voltages_injections = check_power_system_violations(net, print=False)
#             if isinstance(voltage_violations, str):
#                 voltage_violations = pd.DataFrame(columns=['vm_pu', 'va_degree', 'p_mw', 'q_mvar', 'Bus name'])
#             if isinstance(overloaded_lines, str):
#                 overloaded_lines = pd.DataFrame(columns=['line_name', 'line_loading_percent'])
#             if isinstance(overloaded_transformers, str):
#                 overloaded_transformers = pd.DataFrame(columns=['trafo_name', 'trafo_loading_percent'])
#             if isinstance(line_loadings, str):
#                 line_loadings = pd.DataFrame(columns=['line_name', 'line_loading_percent'])
#             if isinstance(trafo_loadings, str):
#                 trafo_loadings = pd.DataFrame(columns=['trafo_name', 'trafo_loading_percent'])
#             if isinstance(bus_voltages_injections, str):
#                 bus_voltages_injections = pd.DataFrame(columns=['bus_name', 'voltage_pu'])

#             voltage_violations_list.append(voltage_violations)
#             overloaded_lines_list.append(overloaded_lines)
#             overloaded_transformers_list.append(overloaded_transformers)
#             line_loadings_list.append(line_loadings)
#             trafo_loadings_list.append(trafo_loadings)
#             bus_voltages_injections_list.append(bus_voltages_injections)
        
#             # Flatten the lists of nested DataFrames and include the datetime
#     df_voltage_violations = pd.concat([df.assign(Outage=outage) for df, outage in zip(voltage_violations_list, outages)], ignore_index=True)
#     df_overloaded_lines = pd.concat([df.assign(Outage=outage) for df, outage in zip(overloaded_lines_list, outages)], ignore_index=True)
#     df_overloaded_transformers = pd.concat([df.assign(Outage=outage) for df, outage in zip(overloaded_transformers_list, outages)], ignore_index=True)
#     df_line_loadings = pd.concat([df.assign(Outage=outage) for df, outage in zip(line_loadings_list, outages)], ignore_index=True)
#     df_trafo_loadings = pd.concat([df.assign(Outage=outage) for df, outage in zip(trafo_loadings_list, outages)], ignore_index=True)
#     df_bus_voltages_injections = pd.concat([df.assign(Outage=outage) for df, outage in zip(bus_voltages_injections_list, outages)], ignore_index=True)
#         # violations = check_violations(results[f'remove_trafo_{i}'])
#         # results[f'remove_trafo_{i}'].violations = violations

#     return outages

def contingency_analysis_with_violations(net):
    """
    Performs a contingency analysis and checks for violations in the power system.
    Returns a list of outages and dataframes containing violations for each scenario.
    """
    # Results storage
    voltage_violations_list = []
    overloaded_lines_list = []
    overloaded_transformers_list = []
    line_loadings_list = []
    trafo_loadings_list = []
    bus_voltages_injections_list = []
    outages = []

    def ensure_dataframe_format(data, columns):
        """Ensures that the data is in DataFrame format with specified columns."""
        if isinstance(data, str):
            data = pd.DataFrame(columns=columns)
        return data

    def process_violations(net, component_type, idx):
        """
        Removes a component, runs the power flow, and checks for violations.
        If violations are detected, they are stored in the respective lists.
        """
        outage_name = f'outage_{component_type}_{idx}'
        net = remove_and_run_powerflow(net, getattr(net, component_type), idx, component_type)
        violations = check_violations(net)
        
        if violations:
            outages.append(outage_name)
            secure, voltage_violations, overloaded_lines, overloaded_transformers, line_loadings, trafo_loadings, bus_voltages_injections = check_power_system_violations(net, print_results=False)
            
            voltage_violations = ensure_dataframe_format(voltage_violations, ['vm_pu', 'va_degree', 'p_mw', 'q_mvar', 'Bus name'])
            overloaded_lines = ensure_dataframe_format(overloaded_lines, ['line_name', 'line_loading_percent'])
            overloaded_transformers = ensure_dataframe_format(overloaded_transformers, ['trafo_name', 'trafo_loading_percent'])
            line_loadings = ensure_dataframe_format(line_loadings, ['line_name', 'line_loading_percent'])
            trafo_loadings = ensure_dataframe_format(trafo_loadings, ['trafo_name', 'trafo_loading_percent'])
            bus_voltages_injections = ensure_dataframe_format(bus_voltages_injections, ['bus_name', 'voltage_pu'])

            voltage_violations_list.append(voltage_violations)
            overloaded_lines_list.append(overloaded_lines)
            overloaded_transformers_list.append(overloaded_transformers)
            line_loadings_list.append(line_loadings)
            trafo_loadings_list.append(trafo_loadings)
            bus_voltages_injections_list.append(bus_voltages_injections)

    # Analyze lines
    for i in net.line.index:
        process_violations(net, 'line', i)

    # Analyze generators
    for i in net.gen.index:
        process_violations(net, 'gen', i)

    # Analyze transformers
    for i in net.trafo.index:
        process_violations(net, 'trafo', i)

    return outages, voltage_violations_list, overloaded_lines_list, overloaded_transformers_list, line_loadings_list, trafo_loadings_list, bus_voltages_injections_list

def save_results_to_excel(scenarios, outages_list, voltage_violations_list, overloaded_lines_list, overloaded_transformers_list, line_loadings_list, trafo_loadings_list, bus_voltages_injections_list, measurement_prod_cons_new_list, excel_path):
    """
    Saves the results of the contingency analysis to an Excel file.
    """
    # Combine results with scenario information
    df_voltage_violations = pd.concat([df.assign(Scenario=scenario, Outage=outage) for scenario, outages, vvl in zip(scenarios, outages_list, voltage_violations_list) for df, outage in zip(vvl, outages)], ignore_index=True)
    df_overloaded_lines = pd.concat([df.assign(Scenario=scenario, Outage=outage) for scenario, outages, oll in zip(scenarios, outages_list, overloaded_lines_list) for df, outage in zip(oll, outages)], ignore_index=True)
    df_overloaded_transformers = pd.concat([df.assign(Scenario=scenario, Outage=outage) for scenario, outages, otl in zip(scenarios, outages_list, overloaded_transformers_list) for df, outage in zip(otl, outages)], ignore_index=True)
    df_line_loadings = pd.concat([df.assign(Scenario=scenario, Outage=outage) for scenario, outages, lll in zip(scenarios, outages_list, line_loadings_list) for df, outage in zip(lll, outages)], ignore_index=True)
    df_trafo_loadings = pd.concat([df.assign(Scenario=scenario, Outage=outage) for scenario, outages, tll in zip(scenarios, outages_list, trafo_loadings_list) for df, outage in zip(tll, outages)], ignore_index=True)
    df_bus_voltages_injections = pd.concat([df.assign(Scenario=scenario, Outage=outage) for scenario, outages, bvil in zip(scenarios, outages_list, bus_voltages_injections_list) for df, outage in zip(bvil, outages)], ignore_index=True)
     # Write consumption profiles to Excel
    df_consumption_profiles = pd.concat([df.assign(Scenario=scenario) for df, scenario in zip(measurement_prod_cons_new_list, scenarios)], ignore_index=True)

    # # Define the path to the Excel file
    # excel_path = 'results_contingency_analysis.xlsx'

    # Write DataFrames to different sheets
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df_consumption_profiles.to_excel(writer, sheet_name='Consumption Profiles', index=False)
        df_voltage_violations.to_excel(writer, sheet_name='Voltage Violations', index=False)
        df_overloaded_lines.to_excel(writer, sheet_name='Overloaded Lines', index=False)
        df_overloaded_transformers.to_excel(writer, sheet_name='Overloaded Transformers', index=False)
        df_line_loadings.to_excel(writer, sheet_name='Line Loadings', index=False)
        df_trafo_loadings.to_excel(writer, sheet_name='Transformer Loadings', index=False)
        df_bus_voltages_injections.to_excel(writer, sheet_name='Bus Voltages Injections', index=False)

    print(f"Results have been saved to {excel_path}")
    from openpyxl import load_workbook
    # Open the workbook to add filters
    workbook = load_workbook(excel_path)

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        # Add a filter to the first row
        sheet.auto_filter.ref = sheet.dimensions
        # Save the workbook with the filters applied
    workbook.save(excel_path)


def contingency_analysis_outages(net):
    # results = []
    outages = []
    # print('net index', net.line.index)

    # Analyze lines
    for i in net.line.index:
        net = remove_and_run_powerflow(net, net.line, i, 'line')
        violations = check_violations(net)
        # print('violations:',violations)
        if violations is True:
            outages.append(f'remove_line_{i}')

            voltage_violations = net.res_bus[(net.res_bus.vm_pu > net.bus.max_vm_pu) | (net.res_bus.vm_pu < net.bus.min_vm_pu)]
            voltage_violations= voltage_violations.drop(28)  # remove ref bus
        # results[f'check_violations(net)
        # # results[f'remove_line_{i}'].violations = violations

    # Analyze generators
    for i in net.gen.index:
        net = remove_and_run_powerflow(net, net.gen, i, 'gen')
        violations = check_violations(net)
        # results[f'remove_gen_{i}'] = remove_and_run_powerflow(net, net.gen, i, 'gen')
        # violations = check_violations(results[f'remove_gen_{i}'])
        if violations is True:
            outages.append(f'remove_gen_{i}')
        # violations = check_violations(results[f'remove_gen_{i}'])
        # results[f'remove_gen_{i}'].violations = violations

    # Analyze transformers
    for i in net.trafo.index:
        # results[f'remove_trafo_{i}'] = remove_and_run_powerflow(net, net.trafo, i, 'trafo')
        # violations = check_violations(results[f'remove_trafo_{i}'])
        net = remove_and_run_powerflow(net, net.trafo, i, 'trafo')
        violations = check_violations(net)
        if violations is True:
            outages.append(f'remove_trafo_{i}')
        # violations = check_violations(results[f'remove_trafo_{i}'])
        # results[f'remove_trafo_{i}'].violations = violations

    return outages


def sample_normal_distributions(input_array, std_dev=0.1, rnd_seed = 10):
    """
    Creates a new array by sampling from normal distributions centered at each point in the input array.
    
    Parameters:
    input_array (numpy.ndarray): Array of input points.
    std_dev (float or numpy.ndarray): Standard deviation(s) for the normal distributions. 
                                      Can be a single float value or an array of the same length as input_array.
    
    Returns:
    numpy.ndarray: Array of sampled points.
    """
    np.random.seed(rnd_seed)
    if isinstance(std_dev, (float, int)):
        std_devs = np.full_like(input_array, std_dev, dtype=float)
    elif isinstance(std_dev, np.ndarray) and std_dev.shape == input_array.shape:
        std_devs = std_dev
    else:
        raise ValueError("std_dev must be a float or a numpy array of the same length as input_array")
    
    # Initialize an empty list to store sampled points
    sampled_points = []
    
    # Loop through each point in the input array and its corresponding standard deviation
    for point, std in zip(input_array, std_devs):
        # Create a normal distribution centered at 'point' with its corresponding standard deviation
        sampled_point = np.random.normal(loc=point, scale=std)
        # Append the sampled point to the list
        sampled_points.append(sampled_point)
    
    # Convert the list to a numpy array
    new_array = np.array(sampled_points)
    
    return new_array

def calculate_reactive_power(active_power, power_factor):
    apparent_power = active_power / power_factor
    reactive_power = np.sqrt(apparent_power**2 - active_power**2)
    return reactive_power
